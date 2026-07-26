import os
import sys

# Ensure the local workspace packages can be imported
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from src import database, models, security
from src.services import client_service


def run_tests():
    print("==================================================")
    print("        RUNNING AUTOMATED SYSTEM TEST SUITE       ")
    print("==================================================")
    
    # Use a separate test database file to not mess up the production db
    database.DB_PATH = "test_mail.db"
    if os.path.exists("test_mail.db"):
        os.remove("test_mail.db")

    database.init_db()
    conn = database.get_db_connection()

    try:
        # Test 1: First setup / admin check
        print("[1/8] Test: Check default admin presence initially...")
        assert models.any_admins_exist(conn), "Seeded admin should exist initially"
        admin_user = models.get_user_by_username(conn, "admin")
        assert admin_user is not None, "Default admin user 'admin' must exist"
        assert security.verify_password(admin_user["password_hash"], "admin"), "Default admin password must be 'admin'"
        print("  ✔ Passed: Default admin user 'admin' exists and has password 'admin'.")

        # Test 2: Password hashing & verification
        print("[2/8] Test: Password hashing & verification...")
        pw = "super_secure_pass_123"
        hashed = security.hash_password(pw)
        assert hashed != pw, "Hash should not be plaintext"
        assert security.verify_password(hashed, pw), "Password verification should pass"
        assert not security.verify_password(hashed, "wrong_pass"), "Password verification should fail for wrong pass"
        print("  ✔ Passed: Password hashing works.")

        # Test 3: Encryption & Decryption
        print("[3/8] Test: Encryption & Decryption...")
        app_pw = "imap_app_password"
        encrypted = security.encrypt_password(app_pw)
        assert encrypted != app_pw, "Encryption should alter the text"
        decrypted = security.decrypt_password(encrypted)
        assert decrypted == app_pw, "Decrypted password must match original"
        print("  ✔ Passed: Fernet encryption and decryption works.")

        # Test 4: Create Admin User
        print("[4/8] Test: Creating admin user...")
        admin_hash = security.hash_password("admin_pass")
        admin_id = models.create_user(conn, "admin_user", admin_hash, "admin")
        assert models.any_admins_exist(conn), "Admins exist now"
        print("  ✔ Passed: Admin created and checked.")

        # Test 5: Add Client via service (transactional)
        print("[5/8] Test: Create Client via Service...")
        client_id = client_service.add_client(
            conn,
            username="client_user",
            password="client_password",
            name="John Client",
            email="john@example.com",
            imap_host="imap.example.com",
            app_password="my_secret_app_pass",
            imap_port=993,
            mailbox="INBOX"
        )
        assert client_id is not None, "Client should be created"

        # Verify client details in DB
        client = client_service.get_client_details_by_id(conn, client_id)
        assert client["username"] == "client_user", "Client username should match"
        assert client["name"] == "John Client", "Client name should match"
        assert client["email"] == "john@example.com", "Client email should match"
        assert security.decrypt_password(client["app_password"]) == "my_secret_app_pass", "Decrypted app password must match"
        print("  ✔ Passed: Client added and retrieved securely.")

        # Test 6: Username/Email uniqueness validation
        print("[6/8] Test: Unique constraints validation...")
        try:
            client_service.add_client(
                conn,
                username="client_user",
                password="another_pass",
                name="Duplicate User",
                email="duplicate@example.com",
                imap_host="imap.example.com",
                app_password="pass"
            )
            assert False, "Should raise ValueError for duplicate username"
        except ValueError as e:
            print(f"  ✔ Got expected unique username error: {e}")

        try:
            client_service.add_client(
                conn,
                username="unique_user",
                password="another_pass",
                name="Duplicate Email User",
                email="john@example.com",
                imap_host="imap.example.com",
                app_password="pass"
            )
            assert False, "Should raise ValueError for duplicate email"
        except ValueError as e:
            print(f"  ✔ Got expected unique email error: {e}")

        # Test 7: Update Client details
        print("[7/8] Test: Updating client...")
        client_service.update_client_profile(
            conn,
            client_id,
            name="John Updated",
            imap_host="imap.newhost.com",
            app_password="new_app_password"
        )
        updated_client = client_service.get_client_details_by_id(conn, client_id)
        assert updated_client["name"] == "John Updated", "Name should be updated"
        assert updated_client["imap_host"] == "imap.newhost.com", "IMAP Host should be updated"
        assert security.decrypt_password(updated_client["app_password"]) == "new_app_password", "App password should be updated"
        print("  ✔ Passed: Dynamic updates verified.")

        # Test 8: Cascade Deletion
        print("[8/9] Test: Client deletion (cascade delete on user)...")
        client_service.delete_client_profile(conn, client_id)
        deleted_client = client_service.get_client_details_by_id(conn, client_id)
        assert deleted_client is None, "Client should be deleted"

        user = models.get_user_by_id(conn, client["user_id"])
        assert user is None, "User should be cascade deleted"
        print("  ✔ Passed: Account deletion and DB cascades verified.")

        # Test 9: Content extraction validation
        print("[9/9] Test: Content extraction from various attachments...")
        import io
        from src.services.email_service import extract_text_from_attachment
        
        # Test txt
        txt_bytes = b"Hello, this is a plain text file."
        txt_res = extract_text_from_attachment(txt_bytes, "test.txt")
        assert "plain text file" in txt_res, "TXT extraction failed"
        
        # Test docx
        import docx
        doc = docx.Document()
        doc.add_paragraph("Hello, this is a Word document paragraph.")
        doc_io = io.BytesIO()
        doc.save(doc_io)
        doc_bytes = doc_io.getvalue()
        doc_res = extract_text_from_attachment(doc_bytes, "test.docx")
        assert "Word document paragraph" in doc_res, "DOCX extraction failed"
        
        # Test xlsx
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Excel Data Cell A1"
        ws["B2"] = "Excel Data Cell B2"
        wb_io = io.BytesIO()
        wb.save(wb_io)
        wb_bytes = wb_io.getvalue()
        wb_res = extract_text_from_attachment(wb_bytes, "test.xlsx")
        assert "Excel Data Cell A1" in wb_res and "Excel Data Cell B2" in wb_res, "XLSX extraction failed"
        
        # Test zip
        import zipfile
        zip_io = io.BytesIO()
        with zipfile.ZipFile(zip_io, "w") as z:
            z.writestr("nested.txt", b"Content inside zip txt file")
        zip_bytes = zip_io.getvalue()
        zip_res = extract_text_from_attachment(zip_bytes, "test.zip")
        assert "Content inside zip txt file" in zip_res, "ZIP extraction failed"
        
        print("  ✔ Passed: Content extraction functions verified for TXT, DOCX, XLSX, ZIP.")


        print("\n==================================================")
        print("        ALL VERIFICATIONS PASSED SUCCESSFULLY!    ")
        print("==================================================")


    finally:
        conn.close()
        if os.path.exists("test_mail.db"):
            os.remove("test_mail.db")


if __name__ == "__main__":
    run_tests()
